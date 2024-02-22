import csv
import inspect
import logging
import softest
from openpyxl import workbook, load_workbook


class Utils(softest.TestCase):  # allstop1 and any value to verify text
    def assert_list_item_text(self, list1, value):
        """

        :param list1: Accept list as argument
        :param value: Value that can be verified
        :return: Assertion is passed else failed
        """
        for stop in list1:
            print("The Text is: " + stop.text)
            # self.soft_assert(self.assertEqual(stop.text, value, "Assertion error"))
            self.soft_assert(self.assertEqual, stop.text, value)
            # this soft assert will verify all the 25 search results even if one assertion failed happen/
            # verification failed happen
            if stop.text == value:
                print("Test is Passed")
            # elif stop.text != "1 Stop":
            #     print("Test is Passed")
            else:
                print("Test is failed")
        # self.assert_all()

    def custom_logger(loglevel=logging.DEBUG):
        # 1 set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # new_logger = logging.getLogger(CustomLoggerDemo.__name__)
        new_logger = logging.getLogger(logger_name)
        # 2
        new_logger.setLevel(logging.DEBUG)
        # 3
        file_handler = logging.FileHandler("automation_logs.log", mode="a")
        # 4
        formatter1 = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s',
                                       datefmt='%m/%d/%y %I:%M:%S :%p')
        # 5
        file_handler.setFormatter(formatter1)
        # 6
        new_logger.addHandler(file_handler)
        # # 7
        return new_logger

    def read_data_from_excel(file_name, sheet):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_count = sh.max_row
        column_count = sh.max_column

        for i in range(2, row_count + 1):
            row = []
            for j in range(1, column_count + 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list

    def read_data_from_csv(filename):
        # Create the Empty list
        datalist = []
        csvdata = open(filename, "r")
        # create the csv reader
        reader = csv.reader(csvdata)
        # Skip Header
        next(reader)
        # Add csv rows in list
        for rows in reader:
            datalist.append(rows)
        return datalist


"""        for stop in list1:
            print("The Text is: ", stop.text)
            try:
                assert stop.text == value
                print("Assertion Pass")
            except:
                print("Assertion is failed")
"""

# Now create the object of this class in the test_caseone.py file
